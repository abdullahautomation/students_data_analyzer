import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Side, Border
def start(input, output):
    df = pd.read_csv(input)
    df["Average"] = ((df["Math"] + df["English"] + df["Science"])/3).round(2)
    results = []
    for avg in df["Average"]:
        if avg >= 50:
            results.append("Pass")
        else:
            results.append("Fail")
    df["Results"] = results
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Result Report"
    headers = ["Name", "Math", "English", "Science", "Average", "Results"]
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append([row["Name"], row["Math"], row["English"], row["Science"], row["Average"], row["Results"]])
    green = PatternFill(fill_type="solid", fgColor="00FF00")
    red = PatternFill(fill_type="solid", fgColor="FF0000")

    for index, row in df.iterrows():
        row_data = index + 2
        if row["Results"] =="Pass":
            ws.cell(row=row_data, column=1).fill = green
            ws.cell(row=row_data, column=2).fill = green
            ws.cell(row=row_data, column=3).fill = green
            ws.cell(row=row_data, column=4).fill = green
            ws.cell(row=row_data, column=5).fill = green
            ws.cell(row=row_data, column=6).fill = green
        else:
            ws.cell(row=row_data, column=1).fill = red
            ws.cell(row=row_data, column=2).fill = red
            ws.cell(row=row_data, column=3).fill = red
            ws.cell(row=row_data, column=4).fill = red
            ws.cell(row=row_data, column=5).fill = red
            ws.cell(row=row_data, column=6).fill = red
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color="000000")
    thin = Side(style="thin")
    bor_in = Border(top=thin, bottom=thin, left=thin, right=thin)
    thick = Side(style="thick")
    bor_ick = Border(top=thick, bottom=thick, left=thick, right=thick)
    for cell in ws[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        cell.border = bor_ick
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = bor_in
    wb.save(output)
    wb.close()
    print("Done! File Saved in Folder. Now you can close this Window.")

start("students.csv", "output_results.xlsx")